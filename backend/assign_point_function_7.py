# encoding: utf-8
import time
import openpyxl
import warnings
import pandas as pd
from pyecharts.charts import *
from pyecharts.charts import Bar
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pyecharts import options as opts
from snapshot_pyppeteer import snapshot
from pyecharts.globals import ThemeType
from pyecharts.render import make_snapshot
from pyecharts.globals import CurrentConfig

warnings.filterwarnings("ignore")


# 赋分导出功能
def assign_point_function(filename, sheet_name="Sheet2"):
    df_data = pd.read_excel(filename, header=2, sheet_name=sheet_name)
    # 获取最低分设置
    df_index = pd.read_excel(filename, header=0, sheet_name=sheet_name)
    columns_numer = df_index.shape[1] + 1
    # 判断第一行是否存在 NaN 值，进行权值索取
    if df_index.iloc[0].isna().any():
        # 使用 iloc[] 选取除第一行以外的所有行和所有列
        # 删除该列
        df_index.dropna(axis=1, inplace=True)
    else:
        # 返回原 DataFrame
        pass
    df_score = df_index.iloc[2:, 1:]
    # 可以使用df.shape属性来查看数据集的行数和列数。其中，df.shape[0]为行数，df.shape[1]为列数
    num_cols = df_score.shape[1]
    num_rows = df_score.shape[0]
    # 重置内存所存在的数，防止出错。然后获取已经设置了的分数
    equal_score = 0
    full_marks = 100
    last_score = []
    for j in range(num_rows):
        score = []
        for i in range(num_cols):
            min_score = df_score.iloc[:, i].min()
            # 获取数据帧中的值并转换为 1D 数组
            max_score = df_score.iloc[:, i].max()
            coefficient_lowest_score = df_index.iloc[0:1, 1:].values.flatten()[i]
            your_score = df_index.iloc[j + 2:j + 3, 1:].values.flatten()[i]
            len_score = max_score - min_score
            other_equal_score = full_marks - coefficient_lowest_score

            equal_score = coefficient_lowest_score + (other_equal_score) * (your_score - min_score) / len_score
            # 如果想要改变生成的浮点数精度，可以从这里结果的修改精度，如equal_score = round(equal_score, 2)保留两位
            equal_score = round(equal_score, 2)
            score.append(equal_score)
        #     求所有项的平均数
        #     result_score=sum(score)/len(score)
        if len(score) > 0:
            result_score = sum(score) / len(score)
        else:
            # 在这里处理除数为零的情况
            result_score = 0
        last_score.append(result_score)
    # print(last_score)
    #  添加新的一列，并且可以设置为整数，下面语句可以把列表转整数
    # last_score = [int(x) for x in last_score]
    df_data['归一化成绩'] = last_score
    #####################################################################################
    # 打开原始工作簿
    workbook = openpyxl.load_workbook(filename)

    # 选择活动工作表
    # worksheet = workbook['Sheet2']
    # 根据sheet名字获得sheet
    sheet = workbook[sheet_name]

    # 获得所有sheet的名称
    print("包含表格：")
    print(workbook.sheetnames)

    # 写入数据到第3行第6列:worksheet.cell(row=3, column=9, value=22)
    # 循环写入列表，从第三行最后一列开始写

    # 插入列表第一个数据为表头
    last_score.insert(0, '归一化成绩')  # 在列表的第一个位置插入'归一化成绩'
    print("列数：" + str(columns_numer))
    for i in range(len(last_score)):
        sheet.cell(row=i + 3, column=columns_numer, value=last_score[i])
        print(last_score[i])
    # 添加筛选器条件到工作表中
    # sheet.auto_filter.ref = "B3:I3"

    # 获取第三行的所有单元格
    row_3 = sheet[3]

    # 添加数据筛选功能
    # sheet.auto_filter.ref = 'B3:' + str(row_3[-1].column) + '3'
    # print(sheet.auto_filter.ref)
    # 如果要删除第二行，解除下面注释功能
    # sheet.delete_rows(2)
    #######################################################################3
    # filename是原文件名，数据格式是字符串，如果不修改的话就不用更改
    save_name = "文件生成//"+"filename.xlsx"
    workbook.save(save_name)
    print("生成归一化成绩成功！")
    return save_name
