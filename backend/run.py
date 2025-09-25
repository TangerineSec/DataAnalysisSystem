import time
import traceback
import urllib
from typing import List

import numpy as np
import pandas as pd
import tornado.ioloop
import tornado.web
# 如何合并两个表
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from tornado_request_mapping import request_mapping, Route
from pyecharts.charts import Bar
import pyecharts.options as opts
from pyecharts.charts import Line
import pyecharts.options as opts
from pyecharts.charts import Line
from pyecharts.globals import ThemeType
import pandas as pd
import numpy as np

pd.set_option("display.max_rows", 500)
pd.set_option("display.max_columns", 500)
from pyecharts.faker import Faker
from pyecharts import options as opts
from pyecharts.charts import Scatter  # 导入散点图绘制模块
from pyecharts.commons.utils import JsCode
from pyecharts.charts import Bar, Grid, Line
from pyecharts.globals import ThemeType
from pyecharts import options as opts
from pyecharts.charts import Bar, Line, Scatter, Page
from bs4 import BeautifulSoup
import pyecharts
import time
from pyecharts import options as opts
from pyecharts.charts import Radar
from pyecharts.globals import CurrentConfig

# 导入自定义函数
from image_test import generator_html_from_excel
from leida_test import bar_chart
from pie_chart_function_6 import pie_chart
from html_rain_class import rain_class
from assign_point_function_7 import assign_point_function
from dimension_radar_chart_function_8 import dimension_radar_chart

print("run后台程序开始运行！")


def picture_view(firstargv, secondargv, thirdargv):
    # 如何一次性只写一个函数就可以增加功能。模范
    df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)
    bar = Bar()
    path = str(int(time.time())) + '_图片9.html'
    bar.render(path=path)
    with open(path, 'r') as rf:
        return rf.read()


def picture_view9(firstargv, secondargv, thirdargv):  # 文件，表格，主题
    table = pd.read_excel(firstargv, header=0, sheet_name=secondargv)
    # 竖向方向有5个年份
    x_number = len(table.index)  # x值5个
    # 横向方向有6个等级
    y_number = len(table.columns[1:])  # y值5个

    x_value = [year + "年" for year in
               table['年份'].astype(str).values.tolist()]  # ["2019年", "2020年", "2021年", "2022年", "2023年"]

    y_value1 = table['不及格'].astype(float).values.tolist()
    y_value2 = table['及格'].astype(float).values.tolist()
    y_value3 = table['中等'].astype(float).values.tolist()
    y_value4 = table['良好'].astype(float).values.tolist()
    y_value5 = table['优秀'].astype(float).values.tolist()

    import pyecharts.options as opts
    from pyecharts.charts import Line
    from pyecharts.globals import ThemeType

    x_data = x_value
    (
        Line(init_opts=opts.InitOpts(
            width="1800px",
            height="800px",
            theme=ThemeType.ESSOS,
            renderer="svg",
            page_title="Stacked Area Charts",
        ))
        .add_xaxis(xaxis_data=x_data)
        .add_yaxis(
            series_name="优秀",
            stack="总量",
            y_axis=y_value5,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
            label_opts=opts.LabelOpts(is_show=True, position="top"),
        )
        .add_yaxis(
            series_name="良好",
            stack="总量",
            y_axis=y_value4,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
            label_opts=opts.LabelOpts(is_show=True),
        )
        .add_yaxis(
            series_name="中等",
            stack="总量",
            y_axis=y_value3,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
            label_opts=opts.LabelOpts(is_show=True),
        )
        .add_yaxis(
            series_name="及格",
            stack="总量",
            y_axis=y_value2,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
            label_opts=opts.LabelOpts(is_show=True),
        )
        .add_yaxis(
            series_name="不及格",
            stack="总量",
            y_axis=y_value1,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
            label_opts=opts.LabelOpts(is_show=True),
        )

        .set_global_opts(
            title_opts=opts.TitleOpts(title="历年得分率等级趋势"),
            tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),
            yaxis_opts=opts.AxisOpts(
                type_="value",
                axistick_opts=opts.AxisTickOpts(is_show=True),
                splitline_opts=opts.SplitLineOpts(is_show=True),
            ),
            xaxis_opts=opts.AxisOpts(type_="category", boundary_gap=False),
        )
        .render("文件生成/stacked_area_chart.html")
    )

    path = '文件生成/stacked_area_chart.html'
    with open(path, 'r') as rf:
        return rf.read()


def picture_view10(firstargv):
    # df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)

    table = pd.read_excel(firstargv)
    name = table[['真实姓名']].values.flatten()
    time_list = table[['视频观看时长']].values.flatten()
    # 转换为十进制小时
    decimal_hours_list = []
    for time_str in time_list:
        hours, minutes, seconds = map(int, time_str.split(':'))
        decimal_hours = hours + minutes / 60 + seconds / 3600
        decimal_hours = round(decimal_hours, 2)
        decimal_hours_list.append(decimal_hours)
    number_video3 = np.array(decimal_hours_list)
    number_spoc = table[['spoc成绩']].values.flatten()
    number_video1 = table[['视频观看个数']].values.flatten()
    number_video2 = table[['视频观看次数']].values.flatten()
    # number_video3=table[['视频观看时长']].values.flatten()
    number_discuss = table[['讨论区评论数+回复数']].values.flatten()
    # 去0平均数计算，可推广
    new_pandas = pd.DataFrame(
        {'Column1': number_spoc, 'Column2': number_video1, 'Column3': number_video2, 'Column4': number_video3,
         'Column5': number_discuss})
    number_discuss = number_discuss[number_discuss != 0]
    avg_spoc = number_spoc.mean().round(2)
    avg_video1 = number_video1.mean().round(2)
    avg_video2 = number_video2.mean().round(2)
    avg_video3 = number_video3.mean().round(2)
    avg_discuss = number_discuss.mean().round(2)
    max_spoc = number_spoc.max()
    max_video1 = number_video1.max()
    max_video2 = number_video2.max()
    max_video3 = number_video3.max()
    max_discuss = number_discuss.max()
    min_spoc = number_spoc.min()
    min_video1 = number_video1.min()
    min_video2 = number_video2.min()
    min_video3 = number_video3.min()
    min_discuss = number_discuss.min()

    for i in range(len(name)):
        list1 = new_pandas.loc[i].values.tolist()
        singlename = name[i]
        data = [{"value": list1, "name": "学员成绩"}]
        valueMax = [
            {"value": [float(max_spoc), float(max_video1), float(max_video2), float(max_video3), float(max_discuss)],
             "name": "最高成绩"}]
        valueAvg = [
            {"value": [float(avg_spoc), float(avg_video1), float(avg_video2), float(avg_video3), float(avg_discuss)],
             "name": "平均成绩"}]
        c_schema = [
            {"name": "spoc成绩", "max": float(100), "min": float(min_spoc)},
            {"name": "视频\n观看个数", "max": float(100), "min": float(min_video1)},
            {"name": "视频\n观看次数", "max": float(100), "min": float(min_video2)},
            {"name": "视频\n观看时长", "max": float(100), "min": float(min_video3)},
            {"name": "讨论区评论\n数+回复数", "max": float(100), "min": float(min_discuss)},
        ]
        c = (
            Radar(init_opts=opts.InitOpts(width="800px",  # 修改默认宽度，这里可以修改宽度显示900px为佳
                                          height="700px",  # 修改默认高度
                                          theme=ThemeType.WESTEROS,
                                          renderer="svg", ))
            .set_colors(["#4587E7"])
            .add_schema(
                schema=c_schema,
                shape="circle",
                center=["50%", "50%"],
                radius="80%",
                angleaxis_opts=opts.AngleAxisOpts(
                    min_=0,
                    max_=360,
                    is_clockwise=False,
                    interval=5,
                    axistick_opts=opts.AxisTickOpts(is_show=False),
                    axislabel_opts=opts.LabelOpts(is_show=False),
                    axisline_opts=opts.AxisLineOpts(is_show=False),
                    splitline_opts=opts.SplitLineOpts(is_show=False),
                ),
                radiusaxis_opts=opts.RadiusAxisOpts(
                    min_=0,
                    max_=100,
                    interval=15,
                    splitarea_opts=opts.SplitAreaOpts(
                        is_show=True, areastyle_opts=opts.AreaStyleOpts(opacity=1)
                    ),
                ),
                polar_opts=opts.PolarOpts(),
                splitarea_opt=opts.SplitAreaOpts(is_show=False),
                splitline_opt=opts.SplitLineOpts(is_show=False),
                textstyle_opts=opts.TextStyleOpts(font_size=23, color="#201e2f", font_family="楷体",
                                                  font_weight="normal"),
            )
            .add(
                series_name="学员成绩",
                data=data,
                areastyle_opts=opts.AreaStyleOpts(opacity=0.4),
                linestyle_opts=opts.LineStyleOpts(type_="dotted", width=1),
            )
            .add(
                series_name="最高成绩",
                data=valueMax,
                color="#f9713c",
                areastyle_opts=opts.AreaStyleOpts(opacity=0.2),
                linestyle_opts=opts.LineStyleOpts(width=3),
            )
            .add(
                series_name="平均成绩",
                data=valueAvg,
                color="#008000",
                areastyle_opts=opts.AreaStyleOpts(opacity=0.3),
                linestyle_opts=opts.LineStyleOpts(type_="dashed", width=3),
            )
            .set_global_opts(
                title_opts=opts.TitleOpts(title="{}用户：SPOC雷达图".format(singlename), pos_left="20%",
                                          title_textstyle_opts=opts.TextStyleOpts(font_size=20,
                                                                                  font_family="Microsoft YaHei",
                                                                                  font_weight="bold")),
                legend_opts=opts.LegendOpts(legend_icon="roundRect", pos_left="52%", orient="horizontal",
                                            textstyle_opts=opts.TextStyleOpts(font_size=18,
                                                                              font_family="Microsoft YaHei",
                                                                              font_weight="bold")),
            )
            .render("SPOC雷达图/{}用户.html".format(singlename))  # 需要提前的当前文件夹下创建“SPOC雷达图”，否则会无法压缩文件，因为不能自主创建。
        )

    import shutil
    def zip_folder(folder_path, output_path):
        shutil.make_archive(output_path, 'zip', folder_path)

    # 使用示例
    folder_path = 'SPOC雷达图'  # 源文件
    output_path = 'SPOC雷达图'  # 压缩为SPOC雷达图.zip
    zip_folder(folder_path, output_path)
    return "完成。"


def picture_view11(firstargv, secondargv, thirdargv):  # 文件名，sheet名，theme名。
    table = pd.read_excel(firstargv)
    name = table.iloc[3, 1]
    time_list = table[['视频观看时长']].values.flatten()[:-1]
    # 转换为十进制小时
    decimal_hours_list = []
    for time_str in time_list:
        hours, minutes, seconds = map(int, time_str.split(':'))
        decimal_hours = hours + minutes / 60 + seconds / 3600
        decimal_hours = round(decimal_hours, 2)
        decimal_hours_list.append(decimal_hours)
    # 去0平均数计算，可推广
    trainee_spoc = table.iloc[0, 1]
    trainee_video1 = table.iloc[0, 2]
    trainee_video2 = table.iloc[0, 3]
    trainee_video3 = table.iloc[0, 4]
    trainee_discuss = decimal_hours_list[0]
    max_spoc = table.iloc[1, 1]
    max_video1 = table.iloc[1, 2]
    max_video2 = table.iloc[1, 3]
    max_video3 = table.iloc[1, 4]
    max_discuss = decimal_hours_list[1]
    avg_spoc = table.iloc[2, 1]
    avg_video1 = table.iloc[2, 2]
    avg_video2 = table.iloc[2, 3]
    avg_video3 = table.iloc[2, 4]
    avg_discuss = decimal_hours_list[2]

    data = [{"value": [float(trainee_spoc), float(trainee_video1), float(trainee_video2), trainee_discuss,
                       float(trainee_discuss)], "name": "学员成绩"}]
    valueMax = [{"value": [float(max_spoc), float(max_video1), float(max_video2), max_discuss, float(max_discuss)],
                 "name": "最高成绩"}]
    valueAvg = [{"value": [float(avg_spoc), float(avg_video1), float(avg_video2), avg_discuss, float(avg_discuss)],
                 "name": "平均成绩"}]
    c_schema = [
        {"name": "spoc成绩", "max": float(100), "min": float(0)},
        {"name": "视频\n观看个数", "max": float(100), "min": float(0)},
        {"name": "视频\n观看次数", "max": float(100), "min": float(0)},
        {"name": "视频\n观看时长", "max": float(100), "min": float(0)},
        {"name": "讨论区评论\n数+回复数", "max": float(100), "min": float(0)},
    ]
    c = (
        Radar(init_opts=opts.InitOpts(width="800px",  # 修改默认宽度，这里可以修改宽度显示900px为佳
                                      height="700px",  # 修改默认高度
                                      theme=thirdargv,
                                      renderer="svg", ))
        .set_colors(["#4587E7"])
        .add_schema(
            schema=c_schema,
            shape="circle",
            center=["50%", "50%"],
            radius="80%",
            angleaxis_opts=opts.AngleAxisOpts(
                min_=0,
                max_=360,
                is_clockwise=False,
                interval=5,
                axistick_opts=opts.AxisTickOpts(is_show=False),
                axislabel_opts=opts.LabelOpts(is_show=False),
                axisline_opts=opts.AxisLineOpts(is_show=False),
                splitline_opts=opts.SplitLineOpts(is_show=False),
            ),
            radiusaxis_opts=opts.RadiusAxisOpts(
                min_=0,
                max_=100,
                interval=15,
                splitarea_opts=opts.SplitAreaOpts(
                    is_show=True, areastyle_opts=opts.AreaStyleOpts(opacity=1)
                ),
            ),
            polar_opts=opts.PolarOpts(),
            splitarea_opt=opts.SplitAreaOpts(is_show=False),
            splitline_opt=opts.SplitLineOpts(is_show=False),
            textstyle_opts=opts.TextStyleOpts(font_size=23, color="#201e2f", font_family="楷体", font_weight="normal"),
        )
        .add(
            series_name="学员成绩",
            data=data,
            areastyle_opts=opts.AreaStyleOpts(opacity=0.4),
            linestyle_opts=opts.LineStyleOpts(type_="dotted", width=1),
        )
        .add(
            series_name="最高成绩",
            data=valueMax,
            color="#f9713c",
            areastyle_opts=opts.AreaStyleOpts(opacity=0.2),
            linestyle_opts=opts.LineStyleOpts(width=3),
        )
        .add(
            series_name="平均成绩",
            data=valueAvg,
            color="#008000",
            areastyle_opts=opts.AreaStyleOpts(opacity=0.3),
            linestyle_opts=opts.LineStyleOpts(type_="dashed", width=3),
        )
        .set_global_opts(
            title_opts=opts.TitleOpts(title="{}用户：SPOC雷达图".format(name), pos_left="20%",
                                      title_textstyle_opts=opts.TextStyleOpts(font_size=20,
                                                                              font_family="Microsoft YaHei",
                                                                              font_weight="bold")),
            legend_opts=opts.LegendOpts(legend_icon="roundRect", pos_left="52%", orient="horizontal",
                                        textstyle_opts=opts.TextStyleOpts(font_size=18, font_family="Microsoft YaHei",
                                                                          font_weight="bold")),
        )
        .render("{}用户.html".format(name))  # 需要提前的当前文件夹下创建“SPOC雷达图”，否则会无法压缩文件，因为不能自主创建。
    )
    with open("{}用户.html".format(name), 'r') as rf:
        return rf.read()


def picture_view12(firstargv, secondargv, thirdargv):
    df = pd.read_excel(firstargv)

    def generation_normalized_score(i):
        print("正在计算第{}列的归一化数据。".format(i))
        df['归一化-{}'.format(i)] = ((df[i] - df[i].min()) / (df[i].max() - df[i].min()) * 40 + 60).round(2)

    table_name = df.columns[2:]
    for j in table_name:
        print(j)
        generation_normalized_score(j)
    len_row = len(table_name)
    df['归一化-平均成绩'] = df.iloc[:, -len_row:].apply(lambda x: x.sum(), axis=1)/len_row
    filename = ".//文件生成//功能12"+str(int(time.time())) + ".xlsx"
    df.to_excel(filename, index=False)
    return filename
    # 返回excel文件


def picture_view13(firstargv, secondargv, thirdargv):
    table = pd.read_excel(firstargv)
    df = table.iloc[:-1, :]

    def generation_normalized_score(i):
        print("正在计算第{}列的归一化数据。".format(i))
        basic_score = table[i].values[-1]
        add_score = 100 - basic_score
        df['归一化-{}'.format(i)] = (
                    (df[i] - df[i].min()) / (df[i].max() - df[i].min()) * add_score + basic_score).round(2)

    table_name = df.columns[2:]
    for j in table_name:
        print(j)
        generation_normalized_score(j)
    len_row = len(table_name)
    df['归一化-平均成绩'] = df.iloc[:, -len_row:].apply(lambda x: x.sum(), axis=1)/len_row
    filename = ".//文件生成//功能13"+str(int(time.time())) + ".xlsx"
    df.to_excel(filename, index=False)
    return filename
    # 返回excel文件


def picture_view14(firstargv, secondargv, thirdargv):
    df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)
    bar = Bar()
    path = str(int(time.time())) + '_图片14.html'
    bar.render(path=path)
    with open(path, 'r') as rf:
        return rf.read()


def picture_view15(firstargv, secondargv, thirdargv):
    df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)
    bar = Bar()
    path = str(int(time.time())) + '_图片14.html'
    bar.render(path=path)
    with open(path, 'r') as rf:
        return rf.read()


def picture_view16(firstargv, secondargv, thirdargv):
    df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)
    bar = Bar()
    path = str(int(time.time())) + '_图片14.html'
    bar.render(path=path)
    with open(path, 'r') as rf:
        return rf.read()


def picture_view17(firstargv, secondargv, thirdargv):
    df = pd.read_excel(firstargv, header=secondargv, sheet_name=thirdargv)
    bar = Bar()
    path = str(int(time.time())) + '_图片14.html'
    bar.render(path=path)
    with open(path, 'r') as rf:
        return rf.read()


def _load_excel_data(filename, sheet_name=None):
    try:
        wb: Workbook = load_workbook(filename, data_only=True)
        if sheet_name is None:
            # sheet_names = wb.get_sheet_names()
            # ws = wb.[sheet_names[0]]
            sheet_names = wb.sheetnames
            ws = wb[sheet_names[0]]
        else:
            ws = wb[sheet_name]
        rows = ws.max_row
        cols = ws.max_column
        headers = []  # 表头
        for c in range(1, cols + 2):
            v = ws.cell(1, c).value
            headers.append({
                'code': str(c),
                'name': v
            })
        data = []  # [{name: value, name1: value1}]
        for r in range(1, rows + 1):
            rows_dict = {}
            for c in range(1, cols + 1):
                v = ws.cell(r, c).value
                rows_dict[str(c)] = v
            data.append(rows_dict)
        return headers, data
    except:
        print("检查输入文件！")


def request_excel_data(filename, sheet_name="Sheet2"):
    wb: Workbook = load_workbook(filename, data_only=True)
    if sheet_name is None:
        # sheet_names = wb.get_sheet_names()
        # ws = wb.[sheet_names[0]]
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
    else:
        ws = wb[sheet_name]
    rows = ws.max_row
    cols = ws.max_column
    headers = []  # 表头
    for c in range(1, cols + 2):
        v = ws.cell(1, c).value
        headers.append({
            'code': str(c),
            'name': v
        })
    data = []  # [{name: value, name1: value1}]
    for r in range(1, rows + 1):
        rows_dict = {}
        for c in range(1, cols + 1):
            v = ws.cell(r, c).value
            rows_dict[str(c)] = v
        data.append(rows_dict)
    print(data)
    return headers, data


def _get_sheet_names(filename) -> List[str]:
    wb: Workbook = load_workbook(filename, data_only=True)
    sheet_names = wb.sheetnames
    return sheet_names


# 调用函数加权
def _excel_pproportion(filename, header=0):
    # 读取文件
    df = pd.read_excel(filename, sheet_name=0, header=header)
    # 将 '缓考' 和 '补考' 和NaN替换为 0
    df.replace(['缓考', '补考', np.nan], 0, inplace=True)
    # 从提取包含数字的列表data，到把结果添加到原来的df
    # 提取pandas DataFrame或Series的最后一行作为列表，并且仅保留大于0的数字：保存一个列表，再提取数字大于0
    # 获取DataFrame的最后一行
    result = list(df.iloc[-1])
    # 可以使用enumerate函数来记录元素的索引值
    result = [(i, x) for i, x in enumerate(result) if isinstance(x, (int, float)) and x > 0]

    number = len(result)
    ############################################################################################333
    # 计算变量，初始必需data设置为0，而不是初始设置为一个序列Series
    data = 0
    for index, value in result:
        data = data + df.iloc[0:-1, index] * float(value)  # 核心表达式
    # 使用 drop() 方法删除最后一行，注意执行一次就会删除最后一行一次，所以只需要执行一次
    df = df.drop(df.index[-1])
    # astype()函数是用来转换数据类型的，下面这里将学号列的数据类型转换为字符串类型
    try:
        df['学号'] = df['学号'].astype(str)
    except:
        pass
    #  添加新的一列
    df['加权成绩'] = data
    df.replace(['缓考', '补考', np.nan], 0, inplace=True)
    #####################################################################################
    # 从保存文件，到设置字体，文件名。
    # 设置sheet名字，文件名，行索引，字体，字体大小
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)

    workbook = load_workbook(filename)
    sheet = workbook['Sheet1']
    # 添加筛选功能
    sheet.auto_filter.ref = sheet.dimensions
    # 打开文件后设置字体
    font = Font(name='等线', size=11)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
    #######################################################################################
    new_file = '加权' + filename
    workbook.save(new_file)
    # print(f"已经生成文件:{new_file}")
    return new_file


# 合并表格
def merge_excel(filename1, filename2, sheet_name=0, header1=0, header2=0):
    # 读入两个excel表格的数据
    df1 = pd.read_excel(filename1, sheet_name=sheet_name, header=header1)
    df2 = pd.read_excel(filename2, sheet_name=sheet_name, header=header2)

    # 指定用于合并的列
    merge_cols = ['姓名', '学号']

    # 合并两个表格
    merged_df = pd.merge(df1, df2, on=merge_cols)
    new_file_name = 'merge_' + filename1 + filename2 + '.xlsx'

    # 将合并后的数据保存到新的excel文件中
    merged_df.to_excel(new_file_name, index=False)
    ##########################################################
    #     文件格式
    workbook = load_workbook(new_file_name)
    sheet = workbook['Sheet1']
    # 添加筛选功能
    sheet.auto_filter.ref = sheet.dimensions
    # 打开文件后设置字体
    font = Font(name='等线', size=11)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
    workbook.save(new_file_name)
    ###########################################################
    return new_file_name


@request_mapping('/api')
class MyHandler(tornado.web.RequestHandler):
    @request_mapping('/upload_and_get_data1', method='post')
    async def upload_and_get_data1(self):
        """上传并且获取表格中的数据"""
        self.set_header('Access-Control-Allow-Origin', '*')
        files = self.request.files.get('file')
        body: bytes = files[0]['body']
        filename = str(int(time.time())) + '.xlsx'
        with open(filename, 'wb') as wf:
            wf.write(body)
        try:
            headers, rows_list = _load_excel_data(filename)
            self.write({
                'code': 200,
                'msg': '',
                'data': {
                    'headers': headers,
                    'data': rows_list,
                    'filename': filename
                }
            })
        except Exception as e:
            print('----------', traceback.format_exc())
            self.write({
                'code': 400,
                'msg': str(e),
            })

    @request_mapping('/upload_and_get_data_sheet', method='post')
    async def upload_and_get_data_sheet(self):
        """上传并且获取表格中的数据"""
        self.set_header('Access-Control-Allow-Origin', '*')
        files = self.request.files.get('file')
        body: bytes = files[0]['body']
        filename = str(int(time.time())) + '.xlsx'
        with open(filename, 'wb') as wf:
            wf.write(body)
        try:
            headers, rows_list = request_excel_data(filename)
            self.write({
                'code': 200,
                'msg': '',
                'data': {
                    'headers': headers,
                    'data': rows_list,
                    'filename': filename
                }
            })
        except Exception as e:
            print('----------', traceback.format_exc())
            self.write({
                'code': 400,
                'msg': str(e),
            })

    @request_mapping('/excel_proportion')
    async def excel_proportion(self):
        """处理加权excel"""
        self.set_header('Access-Control-Allow-Origin', '*')
        filename = self.get_argument('filename')
        # 函数修改
        generate_filename = _excel_pproportion(filename)
        print(generate_filename)
        print(type(generate_filename))
        try:
            headers, data = _load_excel_data(generate_filename)
            self.write({
                'code': 200,
                'msg': '请求成功！',
                'data': {
                    'headers': headers,
                    'data': data,
                    'filename': generate_filename
                }
            })

        except Exception as e:
            print('----------', traceback.format_exc())
            self.write({
                'code': 400,
                'msg': str(e),
            })

    @request_mapping('/excel_assign')
    async def excel_assign(self):
        """处理赋分excel"""
        self.set_header('Access-Control-Allow-Origin', '*')
        filename = self.get_argument('filename')
        generate_filename = assign_point_function(filename)
        print(generate_filename)
        print(type(generate_filename))
        try:
            headers, data = request_excel_data(generate_filename)
            self.write({
                'code': 200,
                'msg': '请求成功！',
                'data': {
                    'headers': headers,
                    'data': data,
                    'filename': generate_filename
                }
            })

        except Exception as e:
            print('----------', traceback.format_exc())
            self.write({
                'code': 400,
                'msg': str(e),
            })

    # 导出文件excel
    @request_mapping('/excel_export')
    async def excel_export(self):
        self.set_header('Access-Control-Allow-Origin', '*')
        filename = self.get_argument('filename')
        with open(filename, 'rb') as rf:
            c = rf.read()
        self.set_header('Content-Disposition', 'attachment; filename=%s; filename*=%s' %
                        (urllib.parse.quote(filename), urllib.parse.quote(filename)))
        self.set_header('Content-Type', 'application/octet-stream')
        self.write(c)
        self.finish()

    @request_mapping('/upload_two_excel_and_preview', method='post')
    async def upload_two_excel_and_preview(self):
        """上传两个并且获取表格中的数据"""
        self.set_header('Access-Control-Allow-Origin', '*')
        files1 = self.request.files.get('file1')
        files2 = self.request.files.get('file2')
        body1: bytes = files1[0]['body']
        body2: bytes = files2[0]['body']
        filename1 = str(int(time.time())) + '_1.xlsx'
        filename2 = str(int(time.time())) + '_2.xlsx'
        with open(filename1, 'wb') as wf:
            wf.write(body1)
        with open(filename2, 'wb') as wf:
            wf.write(body2)
        header1, data1 = _load_excel_data(filename1)
        header2, data2 = _load_excel_data(filename2)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'headers': header1,
                    'data': data1,
                    'filename': filename1,
                },
                'data2': {
                    'headers': header2,
                    'filename': filename2,
                    'data': data2
                },
            }
        })

    @request_mapping('/merge')
    async def merge(self):
        self.set_header('Access-Control-Allow-Origin', '*')
        filename1 = self.get_argument('filename1')
        filename2 = self.get_argument('filename2')
        new_file = merge_excel(filename1, filename2)
        header, data = _load_excel_data(new_file)
        self.write({
            'code': 200,
            'data': {
                'headers': header,
                'data': data,
                'filename': new_file
            }
        })

    @request_mapping('/upload_three_file_and_get_sheet_names', method='post')
    async def upload_three_file_and_get_sheet_names(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        files1 = self.request.files.get('file1')
        files2 = self.request.files.get('file2')
        files3 = self.request.files.get('file3')
        body1: bytes = files1[0]['body']
        body2: bytes = files2[0]['body']
        body3: bytes = files3[0]['body']
        filename1 = str(int(time.time())) + '_折线图1.xlsx'
        filename2 = str(int(time.time())) + '_折线图2.xlsx'
        filename3 = str(int(time.time())) + '_折线图3.xlsx'
        with open(filename1, 'wb') as wf:
            wf.write(body1)
        with open(filename2, 'wb') as wf:
            wf.write(body2)
        with open(filename3, 'wb') as wf:
            wf.write(body3)
        header1, data1 = _load_excel_data(filename1)
        header2, data2 = _load_excel_data(filename2)
        header3, data3 = _load_excel_data(filename3)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'sheet': _get_sheet_names(filename1),
                    'filename': filename1,
                },
                'data2': {
                    'sheet': _get_sheet_names(filename2),
                    'filename': filename2,
                },
                'data3': {
                    'sheet': _get_sheet_names(filename3),
                    'filename': filename3,
                },
            }
        })

    @request_mapping('/upload_file_and_get_sheet_name', method='post')
    async def upload_file_and_get_sheet_name(self):
        try:
            self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
            files1 = self.request.files.get('file1')
            body1: bytes = files1[0]['body']
            filename1 = ".//上传文件//" + str(int(time.time())) + '.xlsx'
            # filename1 = str(int(time.time())) + '.xlsx'
            # filename1="上传表格.xlsx"
            with open(filename1, 'wb') as wf:
                wf.write(body1)
            header1, data1 = _load_excel_data(filename1)
            self.write({
                'code': 200,
                'data': {
                    'data1': {
                        'sheet': _get_sheet_names(filename1),
                        'filename': filename1,
                    },
                }
            })
        except:
            print("检查输入文件格式")

    @request_mapping('/upload_two_file_and_get_sheet_name', method='post')
    async def upload_two_file_and_get_sheet_name(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        files1 = self.request.files.get('file1')
        files2 = self.request.files.get('file2')
        body1: bytes = files1[0]['body']
        body2: bytes = files2[0]['body']
        filename1 = str(int(time.time())) + '_前期雨课堂图.xlsx'
        filename2 = str(int(time.time())) + '_后期雨课堂图.xlsx'
        with open(filename1, 'wb') as wf:
            wf.write(body1)
        with open(filename2, 'wb') as wf:
            wf.write(body2)
        header1, data1 = _load_excel_data(filename1)
        header2, data2 = _load_excel_data(filename2)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'sheet': _get_sheet_names(filename1),
                    'filename': filename1,
                },
                'data2': {
                    'sheet': _get_sheet_names(filename2),
                    'filename': filename2,
                },
            }
        })

    @request_mapping('/get_three_excel_data', method='get')
    async def get_three_excel_data(self):
        """ 预览 """
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        filename2 = self.get_argument('filename2')
        filename3 = self.get_argument('filename3')

        sheet1 = self.get_argument('sheet1')
        sheet2 = self.get_argument('sheet2')
        sheet3 = self.get_argument('sheet3')

        header1, data1 = _load_excel_data(filename1, sheet1)
        header2, data2 = _load_excel_data(filename2, sheet2)
        header3, data3 = _load_excel_data(filename3, sheet3)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'headers': header1,
                    'data': data1,
                    'filename': filename1,
                },
                'data2': {
                    'headers': header2,
                    'filename': filename2,
                    'data': data2
                },
                'data3': {
                    'headers': header3,
                    'filename': filename3,
                    'data': data3
                },
            }
        })

    @request_mapping('/get_one_excel_data', method='get')
    async def get_one_excel_data(self):
        """预览"""
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        header1, data1 = _load_excel_data(filename1, sheet1)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'headers': header1,
                    'data': data1,
                    'filename': filename1,
                },
            }
        })

    @request_mapping('/get_two_excel_data', method='get')
    async def get_two_excel_data(self):
        """预览"""
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        filename2 = self.get_argument('filename2')
        sheet2 = self.get_argument('sheet2')
        header1, data1 = _load_excel_data(filename1, sheet1)
        header2, data2 = _load_excel_data(filename1, sheet2)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'headers': header1,
                    'data': data1,
                    'filename': filename1,
                },
                'data2': {
                    'headers': header2,
                    'data': data2,
                    'filename': filename2,
                },
            }
        })

    @request_mapping('/generator_html')
    async def generator_html(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        filename2 = self.get_argument('filename2')
        filename3 = self.get_argument('filename3')
        sheet1 = self.get_argument('sheet1')
        sheet2 = self.get_argument('sheet2')
        sheet3 = self.get_argument('sheet3')
        theme = self.get_argument('theme')
        output_html = generator_html_from_excel(filename1, filename2, filename3, sheet1, sheet2, sheet3, theme)
        self.write(output_html)
        # headers, data = _load_excel_data(output_file)
        # with open(output_file, 'rb') as rf:
        #     c = rf.read()
        # self.write(c)

    # 柱状图调用得到文件名字符串，然后导出html
    @request_mapping('/export_html')
    async def export_html(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = bar_chart(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    # 导出文件zip
    @request_mapping('/export_zip')
    async def excel_zip(self):
        self.set_header('Access-Control-Allow-Origin', '*')
        filename = self.get_argument('filename')
        with open(filename, 'rb') as rf:
            c = rf.read()
        self.set_header('Content-Disposition', 'attachment; filename=%s; filename*=%s' %
                        (urllib.parse.quote(filename), urllib.parse.quote(filename)))
        self.set_header('Content-Type', 'application/octet-stream')
        self.write(c)
        self.finish()

    # 饼状图调用得到文件名字符串，然后导出html
    @request_mapping('/export_pie')
    async def export_pie(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = pie_chart(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    # 维度图调用得到文件名字符串，然后导出html
    @request_mapping('/export_dimension_radar_chart')
    async def export_dimension_radar_chart(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = dimension_radar_chart(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    # 维度图调用得到文件名字符串，然后导出html,图片9
    @request_mapping('/export_view9')
    async def export_view9(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view9(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        with open(output_html, 'rb') as file:
            self.write(output_html.read())

    # 维度图调用得到文件名字符串，然后导出html，图片10
    @request_mapping('/export_view10')
    async def export_view10(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        picture_view10(filename1)
        with open("SPOC雷达图.zip", 'rb') as rf:
            c = rf.read()
        self.set_header('Content-Disposition', 'attachment; filename=%s; filename*=%s' %
                        (urllib.parse.quote(filename1), urllib.parse.quote(filename1)))
        self.set_header('Content-Type', 'application/octet-stream')
        self.write(c)
        self.finish()

    # 维度图调用得到文件名字符串，然后导出html，图片11
    @request_mapping('/export_view11')
    async def export_view11(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view11(filename1, sheet1, theme)

        # print(output_html) 生成的文件
        self.write(output_html)

    # 维度图调用得到文件名字符串，然后导出html，图片12
    @request_mapping('/export_view12')
    async def export_view12(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view12(filename1, sheet1, theme)
        with open(output_html, 'rb') as rf:
            c = rf.read()
        self.set_header('Content-Disposition', 'attachment; filename=%s; filename*=%s' %
                        (urllib.parse.quote(filename1), urllib.parse.quote(filename1)))
        self.set_header('Content-Type', 'application/octet-stream')
        # print(output_html) 生成的文件
        self.write(c)
        self.finish()

    # 维度图调用得到文件名字符串，然后导出html，图片13
    @request_mapping('/export_view13')
    async def export_view13(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view13(filename1, sheet1, theme)
        with open(output_html, 'rb') as rf:
            c = rf.read()
        self.set_header('Content-Disposition', 'attachment; filename=%s; filename*=%s' %
                        (urllib.parse.quote(filename1), urllib.parse.quote(filename1)))
        self.set_header('Content-Type', 'application/octet-stream')
        # print(output_html) 生成的文件
        self.write(c)
        self.finish()

    # 维度图调用得到文件名字符串，然后导出html，图片14
    @request_mapping('/export_view14')
    async def export_view14(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view14(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    # 维度图调用得到文件名字符串，然后导出html，图片15
    @request_mapping('/export_view15')
    async def export_view15(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        sheet1 = self.get_argument('sheet1')
        theme = self.get_argument('theme')
        output_html = picture_view15(filename1, sheet1, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    # 雨课堂图调用得到文件名字符串，然后导出html
    @request_mapping('/export_rainclass')
    async def export_rainclass(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename1 = self.get_argument('filename1')
        filename2 = self.get_argument('filename2')
        sheet1 = self.get_argument('sheet1')
        sheet2 = self.get_argument('sheet2')
        theme = self.get_argument('theme')
        # output_html = bar_chart(filename1, sheet1, theme)
        output_html = rain_class(filename1, filename2, sheet1, sheet2, theme)
        # print(output_html) 生成的文件
        self.write(output_html)

    @request_mapping('/get_html_file')
    async def get_html_file(self):
        filename = self.get_argument('filename')
        with open(filename, 'r') as rf:
            c = rf.read()
        self.write(c)
        self.finish()

    # # 雷达图功能添加
    # @request_mapping('/new_api', method='post')
    # async def new_api(self):
    #     self.set_header('Access-Control-Allow-Origin', '*')  # 添加跨域问题解决方法
    #     files = self.request.files.get('file')
    #     file_content = files[0]['body']  # 二进制文件数据
    #     filename = 'hello.xlsx'
    #     with open(filename, 'wb') as wf:
    #         wf.write(file_content)
    #     headers, data = _load_excel_data(filename)
    #     self.write({
    #         'code': 200,
    #         'data': {
    #             'headers': headers,
    #             'data': data
    #         }
    #     })

    @request_mapping('/upload_file', method='post')
    async def upload_file(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        files1 = self.request.files.get('file1')
        body1: bytes = files1[0]['body']
        filename1 = str(int(time.time())) + '_柱状图.xlsx'
        with open(filename1, 'wb') as wf:
            wf.write(body1)
        header1, data1 = _load_excel_data(filename1)
        self.write({
            'code': 200,
            'data': {
                'data1': {
                    'sheet': _get_sheet_names(filename1),
                    'filename': filename1,
                },
            }
        })

    @request_mapping('/new_api', method='post')
    async def generator_barchart(self):
        self.set_header('Access-Control-Allow-Origin', '*')  # 跨域
        filename = self.get_argument('filename')
        theme = self.get_argument('theme')
        output_html = bar_chart(filename, theme)
        self.write(output_html)

    @request_mapping('/generator_html')
    async def generator_image(self):
        filename1 = self.get_argument('filename1')
        filename2 = self.get_argument('filename2')
        filename3 = self.get_argument('filename3')
        sheet1 = self.get_argument('sheet1')
        sheet2 = self.get_argument('sheet2')
        sheet3 = self.get_argument('sheet3')
        theme = self.get_argument('theme')
        output_html = generator_html_from_excel(filename1, filename2, filename3, sheet1, sheet2, sheet3, theme)
        self.write(output_html)

    @request_mapping('/generator_bar')
    async def generator_image(self):
        filename1 = self.get_argument('filename1')
        output_html = bar_chart(filename1)
        self.write(output_html)

    @request_mapping('/excel_bar')
    async def excel_bar(self):
        """处理加权excel"""
        self.set_header('Access-Control-Allow-Origin', '*')
        filename = self.get_argument('filename')
        # generate_filename = _excel_pproportion(filename)
        generate_filename = bar_chart(filename)
        # print(generate_filename)
        # print(type(generate_filename))
        # generate_filename = bar_chart(filename)
        try:
            # headers, data = _load_excel_data(generate_filename)
            # headers, rows_list = _load_excel_data(filename)
            self.write({
                'code': 200,
                'msg': '',
                'data': {
                    # 'headers': headers,
                    # 'data': data,
                    'filename': generate_filename
                }
            })
        except Exception as e:
            print('----------', traceback.format_exc())
            self.write({
                'code': 400,
                'msg': str(e),
            })


if __name__ == "__main__":
    app = tornado.web.Application()

    route = Route(app)
    route.register(MyHandler)

    app.listen(8081)  # 监听端口改变。
    tornado.ioloop.IOLoop.current().start()
